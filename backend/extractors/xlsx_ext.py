"""XLSX text extraction using openpyxl.

For spreadsheets we don't dump every cell — we sample to keep token usage
sane:
  - All sheet names
  - For each sheet: header row + first ~30 rows + last ~5 rows
  - Workbook core properties (title, author, etc.)

Emits per-sheet progress via the optional `progress_cb`.
"""
from __future__ import annotations

from pathlib import Path

from . import ExtractionResult, ProgressCB, UserSkippedError

MAX_ROWS_PER_SHEET = 30
MAX_TAIL_ROWS = 5
MAX_COLS_PER_ROW = 30


def _row_to_text(row) -> str:
    cells = []
    for c in row[:MAX_COLS_PER_ROW]:
        v = c.value
        if v is None:
            cells.append("")
        else:
            s = str(v).strip()
            cells.append(s)
    s = " | ".join(cells).strip(" |")
    return s


def _emit(cb: ProgressCB, cur: int, total: int) -> None:
    if cb is None:
        return
    try:
        cb(cur, total, "sheet")
    except UserSkippedError:
        raise
    except Exception:
        pass


def extract_xlsx(path: Path, progress_cb: ProgressCB = None) -> ExtractionResult:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)

    parts: list[str] = []

    # Core properties
    try:
        props = wb.properties
        if props.title:
            parts.append(f"Title (metadata): {props.title}")
        if props.creator:
            parts.append(f"Author (metadata): {props.creator}")
        if props.lastModifiedBy:
            parts.append(f"Last modified by: {props.lastModifiedBy}")
        if props.subject:
            parts.append(f"Subject (metadata): {props.subject}")
        if props.keywords:
            parts.append(f"Keywords (metadata): {props.keywords}")
        if props.created:
            parts.append(f"Created (metadata): {props.created}")
        if props.modified:
            parts.append(f"Modified (metadata): {props.modified}")
    except Exception:
        pass

    sheet_names = wb.sheetnames
    parts.append(f"Sheets: {', '.join(sheet_names)} (count={len(sheet_names)})")

    total_sheets = len(sheet_names)
    _emit(progress_cb, 0, total_sheets)

    for sheet_idx, name in enumerate(sheet_names, 1):
        ws = wb[name]
        parts.append(f"\n--- Sheet: {name} ---")

        # We can't always know max_row in read-only mode, so iterate manually.
        head_rows: list[str] = []
        all_rows_seen = 0
        try:
            for i, row in enumerate(ws.iter_rows(values_only=False)):
                all_rows_seen += 1
                if i < MAX_ROWS_PER_SHEET:
                    line = _row_to_text(row)
                    if line:
                        head_rows.append(line)
                # Safety cap: we only keep the first 30 rows anyway, so
                # there's no value in scanning further. 5000 rows is plenty
                # to discover the true row count for the [... omitted ...]
                # marker.
                if i > 5000:
                    break
        except Exception as e:
            parts.append(f"[error reading sheet: {e}]")
            continue

        if head_rows:
            parts.extend(head_rows)
        if all_rows_seen > MAX_ROWS_PER_SHEET:
            parts.append(f"[... {all_rows_seen - MAX_ROWS_PER_SHEET} more rows omitted ...]")

        _emit(progress_cb, sheet_idx, total_sheets)

    wb.close()

    text = "\n".join(parts).strip()
    return ExtractionResult(text=text, page_count=len(sheet_names))