"""Atomically (re)generate registry.xlsx from the SQLite store.

Strategy:
  - Build the workbook in memory using openpyxl.
  - Write to <target>.tmp first, then os.replace() to <target>.
  - If <target> is currently open in Excel.exe and the rename fails, log and
    keep the .tmp file so user can rename it manually.
"""
from __future__ import annotations

import logging
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schemas import FileRecord

log = logging.getLogger("docregistrar.excel")

SHEET_NAME = "Documents"

COLUMNS: list[tuple[str, int]] = [
    ("ID", 36),
    ("File name", 30),
    ("Relative path", 50),
    ("Repository path", 60),
    ("Repository", 18),
    ("File size (bytes)", 14),
    ("SHA-256", 36),
    ("Is duplicate", 12),
    ("Duplicate group", 36),
    ("Extension", 10),
    ("Page / slide count", 10),
    ("OS created", 20),
    ("OS modified", 20),
    ("Title", 40),
    ("Description", 50),
    ("Summary", 80),
    ("Document date", 14),
    ("Last update date", 14),
    ("Document type", 18),
    ("Language", 12),
    ("Authors", 30),
    ("Version", 10),
    ("Confidentiality", 18),
    ("Persons", 30),
    ("Organizations", 30),
    ("Locations", 25),
    ("Mentioned dates", 25),
    ("Products / technologies", 35),
    ("Key concepts", 35),
    ("Key phrases (top 10)", 50),
    ("Tags", 30),
    ("Geographic scope", 18),
    ("Industry domain", 20),
    ("Quality score", 10),
    ("Used thinking", 10),
    ("Manually edited", 12),
    ("Status", 12),
    ("Error", 30),
    ("Indexed at", 20),
    ("Custom properties", 60),
]


def _format_custom_properties(items) -> str:
    """Render a list of {key,value} as 'k1: v1 | k2: v2 | ...'.

    Pipe characters inside values are escaped as '\\|' so the cell remains
    parseable.
    """
    if not items:
        return ""
    parts: list[str] = []
    for it in items:
        if not isinstance(it, dict):
            # Pydantic KVPair instance
            try:
                k = str(getattr(it, "key", "") or "")
                v = str(getattr(it, "value", "") or "")
            except Exception:
                continue
        else:
            k = str(it.get("key", "") or "")
            v = str(it.get("value", "") or "")
        k = k.strip()
        v = v.strip().replace("|", r"\|")
        if not k and not v:
            continue
        parts.append(f"{k}: {v}" if k else v)
    return " | ".join(parts)


def _join(items: list[str] | None) -> str:
    if not items:
        return ""
    return "; ".join(s for s in items if s)


def _record_to_row(rec: FileRecord, repo_paths: dict[str, str] | None = None) -> list:
    e = rec.extraction
    # Repository is now a column on FileRecord (was previously inside the JSON).
    repo_name = rec.repository or ""
    repo_path = ""
    if repo_name and repo_paths is not None:
        repo_path = repo_paths.get(repo_name, "") or ""
    return [
        rec.id or "",
        rec.file_name,
        rec.relative_path,
        repo_path,
        repo_name,
        rec.file_size,
        rec.sha256,
        ("Yes" if rec.is_duplicate else ""),
        rec.duplicate_group or "",
        rec.extension,
        rec.page_count if rec.page_count is not None else "",
        rec.os_created or "",
        rec.os_modified or "",
        (e.title if e else ""),
        (e.description if e else ""),
        (e.summary if e else ""),
        (e.document_date if e else ""),
        (e.last_update_date if e else ""),
        (e.document_type if e else ""),
        (e.language if e else ""),
        _join(e.authors) if e else "",
        (e.version if e else ""),
        (e.confidentiality if e else ""),
        _join(e.named_entities.persons) if e else "",
        _join(e.named_entities.organizations) if e else "",
        _join(e.named_entities.locations) if e else "",
        _join(e.named_entities.dates) if e else "",
        _join(e.named_entities.products_technologies) if e else "",
        _join(e.key_concepts) if e else "",
        _join(e.key_phrases) if e else "",
        _join(e.tags) if e else "",
        (e.geographic_scope if e else ""),
        (e.industry_domain if e else ""),
        (round(e.quality_score, 3) if e else ""),
        ("Yes" if rec.used_thinking else ""),
        ("Yes" if rec.manually_edited else ""),
        rec.status,
        rec.error or "",
        rec.indexed_at or "",
        _format_custom_properties(e.custom_properties) if e else "",
    ]


def write_registry(records: list[FileRecord], xlsx_path: Path,
                   repo_paths: dict[str, str] | None = None) -> bool:
    """Write/overwrite registry.xlsx atomically. Returns True on success."""
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(SHEET_NAME)
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(vertical="center", horizontal="left", wrap_text=True)

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    body_align = Alignment(vertical="top", wrap_text=True)
    for rec in records:
        row = _record_to_row(rec, repo_paths)
        ws.append(row)

    # Apply alignment to body
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=c).alignment = body_align

    # AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    try:
        wb.save(tmp_path)
    except Exception as e:
        log.error("Failed to write %s: %s", tmp_path, e)
        return False

    # Atomic replace
    try:
        os.replace(tmp_path, xlsx_path)
        log.info("Registry written: %s (%d rows)", xlsx_path, len(records))
        return True
    except PermissionError as e:
        log.warning(
            "Could not replace %s (probably open in Excel). Kept %s. %s",
            xlsx_path, tmp_path, e,
        )
        return False
    except Exception as e:
        log.error("os.replace failed: %s", e)
        return False


def build_registry_bytes(records: list[FileRecord],
                         repo_paths: dict[str, str] | None = None) -> bytes:
    """Build the registry workbook in memory and return its bytes.

    Used by the GET /api/registry.xlsx endpoint so users can download the
    current state at any time, even while processing is running.
    """
    import io

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(SHEET_NAME)
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(vertical="center", horizontal="left", wrap_text=True)

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    body_align = Alignment(vertical="top", wrap_text=True)
    for rec in records:
        ws.append(_record_to_row(rec, repo_paths))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=c).alignment = body_align

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
